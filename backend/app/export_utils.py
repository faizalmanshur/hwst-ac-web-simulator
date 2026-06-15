from __future__ import annotations

from io import BytesIO
from typing import Any
import math

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle, PageBreak
from reportlab.graphics.shapes import Drawing, Rect, String, Line, Circle, Ellipse, Path


def build_excel(results: dict[str, Any]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ringkasan"
    ws.append(["Output", "Value", "Unit"])
    for row in results.get("summary_rows", []):
        ws.append([row.get("output"), row.get("value"), row.get("unit")])

    ws2 = wb.create_sheet("Data Waktu")
    ts = results.get("time_series", [])
    if ts:
        headers = list(ts[0].keys())
        ws2.append(headers)
        for item in ts:
            ws2.append([item.get(h) for h in headers])

    ws3 = wb.create_sheet("Visualisasi Series")
    vs = results.get("visualization_series", [])
    if vs:
        headers = list(vs[0].keys())
        ws3.append(headers)
        for item in vs:
            ws3.append([item.get(h) for h in headers])

    ws4 = wb.create_sheet("Variasi Volume")
    vv = results.get("volume_variation", [])
    if vv:
        headers = list(vv[0].keys())
        ws4.append(headers)
        for item in vv:
            ws4.append([item.get(h) for h in headers])

    ws5 = wb.create_sheet("Geometri")
    ws5.append(["Komponen", "Parameter", "Value", "Unit"])
    for row in results.get("geometry_rows", []):
        ws5.append([row.get("komponen"), row.get("parameter"), row.get("value"), row.get("unit")])

    ws6 = wb.create_sheet("Persentase Coil")
    coil = results.get("coil_usage", [])
    if coil:
        headers = ["komponen", "zona", "persentase_COP_best_DSH", "persentase_akhir", "t_COP_best_DSH_min", "T_tank_COP_best_DSH_C", "catatan"]
        ws6.append(headers)
        for item in coil:
            ws6.append([item.get(h) for h in headers])

    ph = results.get("ph_series", [])
    if ph:
        ws7 = wb.create_sheet("PH_State")
        ws7.append(["time_min", "label", "name", "h_kJ_kg", "P_kPa", "T_C"])
        for frame in ph:
            for pt in frame.get("points", []):
                ws7.append([frame.get("time_min"), pt.get("label"), pt.get("name"), pt.get("h_kJ_kg"), pt.get("P_kPa"), pt.get("T_C")])

    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 12
            letter = col[0].column_letter
            for cell in col:
                max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
            sheet.column_dimensions[letter].width = min(max_len + 2, 42)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _f(value: Any, nd: int = 2) -> str:
    try:
        return f"{float(value):.{nd}f}"
    except Exception:
        return "-"


def _basic_table_style(font_size: int = 8) -> TableStyle:
    return TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a8a")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.45, colors.HexColor("#94a3b8")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE", (0, 0), (-1, -1), font_size),
    ])


def _latest_row(results: dict[str, Any]) -> dict[str, Any]:
    ts = results.get("visualization_series") or results.get("time_series") or []
    return ts[min(len(ts)-1, max(0, len(ts)//2))] if ts else {}


def _temp_color(t: Any):
    try:
        v = max(0.0, min(1.0, float(t) / 90.0))
    except Exception:
        v = 0.5
    if v < 0.25:
        return colors.HexColor("#0891b2")
    if v < 0.5:
        return colors.HexColor("#22c55e")
    if v < 0.75:
        return colors.HexColor("#f59e0b")
    return colors.HexColor("#dc2626")


def _build_system_drawing(results: dict[str, Any]) -> Drawing:
    row = _latest_row(results)
    d = Drawing(520, 250)
    d.add(Rect(0, 0, 520, 250, rx=14, ry=14, fillColor=colors.HexColor("#f8fafc"), strokeColor=colors.HexColor("#cbd5e1")))
    comp=(55,155); tank=(210,164); cond=(420,165); cap=(340,75); evap=(145,75)
    def pipe(points, temp_key):
        p = Path(); p.moveTo(points[0][0], points[0][1])
        for x,y in points[1:]: p.lineTo(x,y)
        p.strokeColor = _temp_color(row.get(temp_key)); p.strokeWidth = 8; p.fillColor = None
        d.add(p)
    pipe([(comp[0]+42, comp[1]), (tank[0]-92, tank[1])], "T_comp_out_C")
    pipe([(tank[0]+105, tank[1]), (cond[0]-55, cond[1])], "T_ref_out_HWST_C")
    pipe([(cond[0], cond[1]-45), (cond[0], cap[1]+10), (cap[0]+40, cap[1])], "T_cond_out_C")
    pipe([(cap[0]-40, cap[1]), (evap[0]+70, evap[1])], "T_evap_in_C")
    pipe([(evap[0]-70, evap[1]), (comp[0]-20, evap[1]), (comp[0]-20, comp[1]-35)], "T_evap_out_C")
    d.add(Ellipse(comp[0], comp[1], 35, 50, fillColor=colors.white, strokeColor=colors.HexColor("#1e293b"), strokeWidth=1.5))
    d.add(String(comp[0]-28, comp[1]-72, "KOMPRESOR", fontSize=8, fillColor=colors.HexColor("#0f172a")))
    d.add(Rect(tank[0]-90, tank[1]-30, 180, 60, rx=28, ry=28, fillColor=colors.white, strokeColor=colors.HexColor("#1e293b"), strokeWidth=1.5))
    for i in range(8):
        d.add(Line(tank[0]-65+i*17, tank[1]-18, tank[0]-65+i*17, tank[1]+18, strokeColor=colors.HexColor("#f59e0b"), strokeWidth=2.2))
    d.add(String(tank[0]-20, tank[1]-52, "HWST horizontal", fontSize=8, fillColor=colors.HexColor("#0f172a")))
    d.add(Rect(cond[0]-45, cond[1]-50, 90, 100, rx=10, ry=10, fillColor=colors.white, strokeColor=colors.HexColor("#1e293b"), strokeWidth=1.5))
    for i in range(7):
        d.add(Line(cond[0]-30+i*10, cond[1]-35, cond[0]-30+i*10, cond[1]+35, strokeColor=colors.HexColor("#2563eb"), strokeWidth=2))
    d.add(String(cond[0]-28, cond[1]-70, "KONDENSOR", fontSize=8, fillColor=colors.HexColor("#0f172a")))
    d.add(Rect(evap[0]-65, evap[1]-30, 130, 60, rx=10, ry=10, fillColor=colors.white, strokeColor=colors.HexColor("#1e293b"), strokeWidth=1.5))
    for i in range(5):
        d.add(Line(evap[0]-48, evap[1]-18+i*8, evap[0]+48, evap[1]-18+i*8, strokeColor=colors.HexColor("#2563eb"), strokeWidth=2.2))
    d.add(String(evap[0]-36, evap[1]-48, "EVAPORATOR", fontSize=8, fillColor=colors.HexColor("#0f172a")))
    d.add(Circle(cap[0], cap[1], 13, fillColor=colors.HexColor("#f59e0b"), strokeColor=colors.HexColor("#92400e"), strokeWidth=1.5))
    d.add(String(cap[0]-20, cap[1]-30, "KAPILER", fontSize=8, fillColor=colors.HexColor("#0f172a")))
    d.add(String(25, 225, f"T kompresor out: {_f(row.get('T_comp_out_C'),1)} °C", fontSize=8, fillColor=colors.HexColor("#dc2626")))
    d.add(String(25, 211, f"T keluar HWST: {_f(row.get('T_ref_out_HWST_C'),1)} °C", fontSize=8, fillColor=colors.HexColor("#ea580c")))
    d.add(String(25, 197, f"T air tangki: {_f(row.get('T_tank_mean_C'),1)} °C | Q HWST: {_f(row.get('Q_HWST_kW'),2)} kW", fontSize=8, fillColor=colors.HexColor("#334155")))
    return d


def _build_ph_drawing(results: dict[str, Any]) -> Drawing:
    frames = results.get("ph_series", [])
    frame = frames[min(len(frames)-1, max(0, len(frames)//2))] if frames else {"points": []}
    dome = results.get("ph_dome", {}) or {}
    pts = []
    pts += dome.get("liquid", [])[:]
    pts += dome.get("vapor", [])[:]
    pts += frame.get("points", [])
    vals = [(float(p.get("h_kJ_kg")), float(p.get("P_kPa"))) for p in pts if p.get("h_kJ_kg") is not None and p.get("P_kPa")]
    d = Drawing(520, 290)
    d.add(Rect(0, 0, 520, 290, rx=14, ry=14, fillColor=colors.HexColor("#f8fafc"), strokeColor=colors.HexColor("#cbd5e1")))
    if not vals:
        d.add(String(180, 145, "Data P-h belum tersedia", fontSize=12)); return d
    hmin=min(v[0] for v in vals)-25; hmax=max(v[0] for v in vals)+30
    pmin=max(50,min(v[1] for v in vals)*0.75); pmax=max(v[1] for v in vals)*1.25
    lmin=math.log10(pmin); lmax=math.log10(pmax)
    x0,y0,w,h = 55,45,420,205
    def sx(hv): return x0 + (float(hv)-hmin)/max(hmax-hmin,1e-9)*w
    def sy(pv): return y0 + (math.log10(max(float(pv),1))-lmin)/max(lmax-lmin,1e-9)*h
    d.add(Line(x0,y0,x0,y0+h,strokeColor=colors.HexColor("#64748b")))
    d.add(Line(x0,y0,x0+w,y0,strokeColor=colors.HexColor("#64748b")))
    def polyline(points, col, sw=2):
        valid=[p for p in points if p.get("h_kJ_kg") is not None and p.get("P_kPa")]
        if len(valid)<2: return
        path=Path(); path.moveTo(sx(valid[0]["h_kJ_kg"]), sy(valid[0]["P_kPa"]))
        for p in valid[1:]: path.lineTo(sx(p["h_kJ_kg"]), sy(p["P_kPa"]))
        path.strokeColor=col; path.strokeWidth=sw; path.fillColor=None; d.add(path)
    polyline(dome.get("liquid", []), colors.HexColor("#0ea5e9"), 2)
    polyline(dome.get("vapor", []), colors.HexColor("#0f766e"), 2)
    cyc=frame.get("points", [])
    if len(cyc)>1: polyline(cyc+[cyc[0]], colors.HexColor("#1d4ed8"), 3)
    for p in cyc:
        x=sx(p.get("h_kJ_kg")); y=sy(p.get("P_kPa"))
        d.add(Circle(x,y,4.5,fillColor=colors.HexColor("#1d4ed8"),strokeColor=colors.white,strokeWidth=1))
        d.add(String(x+5,y+6,str(p.get("label","")),fontSize=8,fillColor=colors.HexColor("#0f172a")))
    d.add(String(205, 20, "Entalpi, h (kJ/kg)", fontSize=9, fillColor=colors.HexColor("#334155")))
    d.add(String(8, 150, "P log", fontSize=9, fillColor=colors.HexColor("#334155")))
    d.add(String(58, 265, f"Diagram P-h ringkas | t = {_f(frame.get('time_min'),2)} menit", fontSize=10, fillColor=colors.HexColor("#0f172a")))
    return d



def _line_plot_drawing(results: dict[str, Any], title: str, keys: list[str], labels: list[str], units: str = "") -> Drawing:
    ts = results.get("time_series", []) or results.get("visualization_series", []) or []
    d = Drawing(520, 250)
    d.add(Rect(0, 0, 520, 250, rx=14, ry=14, fillColor=colors.HexColor("#f8fafc"), strokeColor=colors.HexColor("#cbd5e1")))
    d.add(String(18, 226, title, fontSize=11, fillColor=colors.HexColor("#0f172a")))
    if not ts:
        d.add(String(200, 120, "Data waktu belum tersedia", fontSize=10, fillColor=colors.HexColor("#64748b")))
        return d
    x0, y0, w, h = 52, 42, 430, 160
    vals: list[float] = []
    for row in ts:
        for key in keys:
            v = row.get(key)
            if isinstance(v, (int, float)) and math.isfinite(float(v)):
                vals.append(float(v))
    if not vals:
        return d
    ymin, ymax = min(vals), max(vals)
    if abs(ymax - ymin) < 1e-9:
        ymin -= 1; ymax += 1
    tmin = float(ts[0].get("time_min", 0) or 0); tmax = float(ts[-1].get("time_min", 1) or 1)
    def sx(t): return x0 + (float(t) - tmin) / max(tmax - tmin, 1e-9) * w
    def sy(v): return y0 + (float(v) - ymin) / max(ymax - ymin, 1e-9) * h
    # axes/grid
    for i in range(5):
        y = y0 + i*h/4
        d.add(Line(x0, y, x0+w, y, strokeColor=colors.HexColor("#e2e8f0"), strokeWidth=0.6))
    d.add(Line(x0, y0, x0, y0+h, strokeColor=colors.HexColor("#94a3b8"), strokeWidth=1))
    d.add(Line(x0, y0, x0+w, y0, strokeColor=colors.HexColor("#94a3b8"), strokeWidth=1))
    palette = ["#2563eb", "#dc2626", "#f59e0b", "#22c55e", "#06b6d4", "#9333ea"]
    for ki, key in enumerate(keys):
        pts = [(sx(row.get("time_min", 0) or 0), sy(row.get(key, 0) or 0)) for row in ts]
        if len(pts) < 2: continue
        path = Path(); path.moveTo(pts[0][0], pts[0][1])
        for x, y in pts[1:]: path.lineTo(x, y)
        path.strokeColor = colors.HexColor(palette[ki % len(palette)]); path.strokeWidth = 2; path.fillColor = None
        d.add(path)
        lx = 315 + (ki % 2)*95; ly = 226 - (ki//2)*13
        d.add(Line(lx, ly, lx+14, ly, strokeColor=colors.HexColor(palette[ki % len(palette)]), strokeWidth=2))
        d.add(String(lx+18, ly-4, labels[ki] if ki < len(labels) else key, fontSize=7, fillColor=colors.HexColor("#334155")))
    d.add(String(18, 22, f"min={ymin:.2f} {units} | max={ymax:.2f} {units}", fontSize=8, fillColor=colors.HexColor("#64748b")))
    d.add(String(410, 22, f"0–{tmax:.1f} min", fontSize=8, fillColor=colors.HexColor("#64748b")))
    return d


def _build_coil_donut_drawing(results: dict[str, Any]) -> Drawing:
    rows = results.get("coil_usage", []) or []
    d = Drawing(520, 240)
    d.add(Rect(0, 0, 520, 240, rx=14, ry=14, fillColor=colors.HexColor("#f8fafc"), strokeColor=colors.HexColor("#cbd5e1")))
    d.add(String(18, 218, "Distribusi Zona Coil pada COP Best DSH", fontSize=11, fillColor=colors.HexColor("#0f172a")))
    color_map = {"DSH": "#dc2626", "TP": "#f59e0b", "SC": "#22c55e", "SH": "#60a5fa"}
    groups = ["HWST", "Kondensor", "Evaporator"]
    def wedge(cx, cy, r, a0, a1, col):
        # crude filled sector polygon
        pts = [(cx, cy)]
        steps = max(4, int(abs(a1-a0)/10))
        for i in range(steps+1):
            a = math.radians(a0 + (a1-a0)*i/steps)
            pts.append((cx + r*math.cos(a), cy + r*math.sin(a)))
        path = Path(); path.moveTo(pts[0][0], pts[0][1])
        for x,y in pts[1:]: path.lineTo(x,y)
        path.closePath()
        path.fillColor = colors.HexColor(col); path.strokeColor = colors.white; path.strokeWidth = 0.8
        d.add(path)
    for gi, g in enumerate(groups):
        cx, cy = 92 + gi*170, 128
        r, rin = 52, 26
        gr = [row for row in rows if row.get("komponen") == g]
        angle = 90
        total = sum(float(row.get("persentase_COP_best_DSH") or 0) for row in gr) or 100.0
        for row in gr:
            val = float(row.get("persentase_COP_best_DSH") or 0)
            if val <= 0: continue
            a1 = angle - 360*val/total
            wedge(cx, cy, r, angle, a1, color_map.get(str(row.get("zona")), "#64748b"))
            angle = a1
        d.add(Circle(cx, cy, rin, fillColor=colors.white, strokeColor=colors.HexColor("#e2e8f0"), strokeWidth=1))
        d.add(String(cx-22, cy-4, g, fontSize=9, fillColor=colors.HexColor("#0f172a")))
        d.add(String(cx-34, 50, ", ".join([f"{row.get('zona')} {float(row.get('persentase_COP_best_DSH') or 0):.1f}%" for row in gr]), fontSize=7, fillColor=colors.HexColor("#334155")))
    return d

def build_pdf(results: dict[str, Any]) -> BytesIO:
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=34, leftMargin=34, topMargin=34, bottomMargin=34)
    styles = getSampleStyleSheet()
    story: list[Any] = []

    story.append(Paragraph("Laporan Simulasi AC + HWST", styles["Title"]))
    story.append(Paragraph("Versi web V9 - sinkron model MATLAB terbaru: compressor calibration, HWST horizontal, visualisasi dinamis, P-h, coil zone, kapiler, dan pressure drop.", styles["Normal"]))
    story.append(Spacer(1, 12))

    data = [["Output", "Value", "Unit"]]
    for row in results.get("summary_rows", []):
        data.append([str(row.get("output", "")), str(row.get("value", "")), str(row.get("unit", ""))])
    table = Table(data, colWidths=[230, 130, 70], repeatRows=1)
    table.setStyle(_basic_table_style(font_size=7))
    story.append(table)

    story.append(PageBreak())
    story.append(Paragraph("Perbandingan dan Distribusi Zona Coil", styles["Heading2"]))
    coil = results.get("coil_usage", [])
    if coil:
        story.append(_build_coil_donut_drawing(results))
        story.append(Spacer(1, 10))
        coil_data = [["Komponen", "Zona", "COP Best DSH (%)", "Akhir (%)", "Catatan"]]
        for row in coil:
            coil_data.append([str(row.get("komponen", "")), str(row.get("zona", "")), str(row.get("persentase_COP_best_DSH", "")), str(row.get("persentase_akhir", "")), str(row.get("catatan", ""))])
        coil_table = Table(coil_data, colWidths=[80, 55, 90, 65, 180], repeatRows=1)
        coil_table.setStyle(_basic_table_style(font_size=8))
        story.append(coil_table)

    story.append(PageBreak())
    story.append(Paragraph("Visualisasi Sistem AC + HWST", styles["Heading2"]))
    story.append(Paragraph("Skema teknis ringkas: kompresor → HWST horizontal → kondensor → kapiler → evaporator → kompresor. Warna pipa mengikuti temperatur timestep tengah/aktif pada data export.", styles["Normal"]))
    story.append(Spacer(1, 10))
    story.append(_build_system_drawing(results))

    story.append(PageBreak())
    story.append(Paragraph("Tren Utama Simulasi", styles["Heading2"]))
    story.append(_line_plot_drawing(results, "Temperatur Titik 1, 2, 2', 3, 4", ["T_evap_out_C","T_comp_out_C","T_ref_out_HWST_C","T_cond_out_C","T_evap_in_C"], ["1 evap out","2 comp out","2' HWST out","3 cond out","4 evap in"], "°C"))
    story.append(Spacer(1, 10))
    story.append(_line_plot_drawing(results, "Suhu Air Tangki dan COP", ["T_tank_mean_C","COP_AC","COP_useful"], ["T tank","COP AC","COP useful"], ""))

    if results.get("ph_series"):
        story.append(PageBreak())
        story.append(Paragraph("Diagram P-h dan State Refrigeran", styles["Heading2"]))
        story.append(Paragraph("Titik 2' menunjukkan refrigeran keluar HWST sebelum masuk kondensor.", styles["Normal"]))
        story.append(Spacer(1, 10))
        story.append(_build_ph_drawing(results))
        story.append(Spacer(1, 10))
        frames = results.get("ph_series", [])
        frame = frames[min(len(frames)-1, max(0, len(frames)//2))]
        ph_data = [["Titik", "Nama", "h (kJ/kg)", "P (kPa)", "T (°C)"]]
        for pt in frame.get("points", []):
            ph_data.append([str(pt.get("label", "")), str(pt.get("name", "")), _f(pt.get("h_kJ_kg"), 2), _f(pt.get("P_kPa"), 1), _f(pt.get("T_C"), 1)])
        ph_table = Table(ph_data, colWidths=[45, 180, 80, 75, 60])
        ph_table.setStyle(_basic_table_style(font_size=8))
        story.append(ph_table)

    story.append(PageBreak())
    story.append(Paragraph("Diagnostik dan Catatan Model", styles["Heading2"]))
    analysis = results.get("analysis_rows", [])
    if analysis:
        headers = list(analysis[0].keys())
        data = [headers] + [[str(row.get(h, "")) for h in headers] for row in analysis]
        tab = Table(data, colWidths=[120, 65, 120, 165], repeatRows=1)
        tab.setStyle(_basic_table_style(font_size=7))
        story.append(tab)
    story.append(Spacer(1, 12))
    story.append(Paragraph("Model V9 mengikuti revisi MATLAB terbaru: CoolProp, kompresor nameplate dengan kalibrasi T_comp_out nominal, Auto-U thermal resistance, zona coil, kapiler/pressure feedback, mixed/lumped tank, dan visualisasi sistem HWST horizontal.", styles["Normal"]))

    doc.build(story)
    buffer.seek(0)
    return buffer
